import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook


def find_fifth_nested_table(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        tables = soup.find_all('table')
        if len(tables) >= 6:
            fifth_table_html = str(tables[6])
            return fifth_table_html
        else:
            print("The page doesn't have at least 6 tables.")
    else:
        print(f"Failed to fetch the URL. Status Code: {response.status_code}")



def extract_links_to_excel(html, filename):
    # Create a BeautifulSoup object from the HTML content
    soup = BeautifulSoup(html, 'html.parser')

    # Create an empty list to store the links
    links = []

    # Find all the <a> tags in the HTML and extract the href attribute
    for a in soup.find_all('a'):
        href = a.get('href')
        if href is not None:
            links.append(href)

    # Load an existing Excel workbook
    wb = load_workbook(filename)

    # Get the active worksheet in the workbook
    sheet = wb.active

    # Find the first empty row in the worksheet
    row_num = 1
    while sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1

    # Write the links to the worksheet
    for i, link in enumerate(links):
        sheet.cell(row=row_num+i, column=1, value=link)

    # Save the modified workbook to disk
    wb.save(filename)


def process_links():
    # Load the links from links.xlsx
    wb = openpyxl.load_workbook('main_links.xlsx')
    sheet = wb.active

    # Iterate through all the filled rows that contain URLs
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        url = row[0]
        if url is not None:
            # Call find_fifth_nested_table() with the URL
            table = find_fifth_nested_table(url)

            # Extract the links from the table and save to link2.xlsx
            extract_links_to_excel(table, 'email_links.xlsx')

process_links()