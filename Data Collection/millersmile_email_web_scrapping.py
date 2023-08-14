import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import openpyxl
import re

def remove_first_n_rows(table_html, n):
    soup = BeautifulSoup(table_html, 'html.parser')
    rows = soup.find_all('tr')
    for i in range(min(n, len(rows))):
        rows[i].extract()
    return str(soup)

def find_fifth_nested_table(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        tables = soup.find_all('table')
        if len(tables) >= 6:
            fifth_table_html = str(tables[6])
            modified_table_html = remove_first_n_rows(fifth_table_html, 7)
            return modified_table_html
        else:
            print("The page doesn't have at least 6 tables.")
    else:
        print(f"Failed to fetch the URL. Status Code: {response.status_code}")

def remove_a_and_img_tags(element):
    for tag in element.find_all(['a', 'img']):
        tag.extract()

def extract_rows_with_strong_tag(html_code):
    soup = BeautifulSoup(html_code, 'html.parser')
    table = soup.find('table')

    if table:
        rows = table.find_all('tr')
        for row in rows:
            td_tags = row.find_all('td')
            for td_tag in td_tags:
                remove_a_and_img_tags(td_tag)
            strong_tags = row.find_all('strong')
            if not strong_tags:
                row.extract()

    return str(soup)

def remove_right_aligned_tds(html):
    soup = BeautifulSoup(html, 'html.parser')
    for td in soup.find_all('td'):
        if td.has_attr('align') and td['align'].lower() == 'right':
            td.extract()
        elif 'Website:' in td.text:
            td.extract()
        else:
            for sub_element in td.descendants:
                if isinstance(sub_element, dict) and 'align' in sub_element:
                    if sub_element['align'].lower() == 'right':
                        td.extract()
    return str(soup)

def remove_control_characters(input_string):
    control_characters = ''.join(map(chr, range(0, 32))) + chr(127)
    control_character_re = '[' + re.escape(control_characters) + ']'
    return re.sub(control_character_re, '', input_string)


def add_tds_to_excel(html, filename):
    soup = BeautifulSoup(html, 'html.parser')
    wb = load_workbook(filename)
    sheet = wb.active

    # Find the first empty row in the worksheet
    row_num = 1
    while sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1

    # Extract the text content of the td tags in each row and add to the worksheet
    row_data = []
    for tr in soup.find_all('tr'):
        for td in tr.find_all('td'):
            cleaned_string = remove_control_characters(td.text.strip())
            row_data.append(cleaned_string)

    sheet.append(row_data)

    # Save the modified workbook
    wb.save(filename)


def process_links():
    # Load the links from links.xlsx
    wb = openpyxl.load_workbook('email_links.xlsx')
    sheet = wb.active

    # Iterate through all the filled rows that contain URLs
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        url = row[0]
        if url is not None:
            # Call find_fifth_nested_table() with the URL
            fifth_table = find_fifth_nested_table(url)
            modified_html = extract_rows_with_strong_tag(fifth_table)
            final_html = remove_right_aligned_tds(modified_html)
            add_tds_to_excel(final_html, 'millersmile_data.xlsx')

process_links()